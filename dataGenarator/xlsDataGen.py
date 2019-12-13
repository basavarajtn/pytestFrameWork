import openpyxl
import pytest

def xlsdata():
    wb=openpyxl.load_workbook("./Data/registorationData.xlsx")
    sh=wb['Sheet1']
    maxr=sh.max_row
    li=[]
    li1=[]
    for i in range(1,maxr+1):
        li1=[]
        un=sh.cell(i,1)
        pw=sh.cell(i,2)
        li1.insert(0,un.value)
        li1.insert(1,pw.value)
        li.insert(i-1,li1)

    print(li)
    return li
